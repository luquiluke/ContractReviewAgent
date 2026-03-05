from io import BytesIO

import openpyxl
import pytest

from app.exporters import build_excel, build_pdf


def test_build_excel_structure():
    buf = build_excel([{"section": "S1", "summary": "Sum1"}])
    assert isinstance(buf, BytesIO)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    assert ws["A1"].value == "Section"
    assert ws["A1"].font.bold is True
    assert ws["B1"].value == "Summary"
    assert ws["B1"].font.bold is True


def test_build_excel_row_count():
    rows = [
        {"section": "S1", "summary": "Sum1"},
        {"section": "S2", "summary": "Sum2"},
        {"section": "S3", "summary": "Sum3"},
    ]
    buf = build_excel(rows)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    assert ws.max_row == 4  # 1 header + 3 data rows


def test_build_pdf_returns_bytes():
    result = build_pdf([{"section": "S1", "summary": "Text"}], "NDA / Confidentiality")
    assert isinstance(result, bytes)
    assert len(result) > 0
    assert result[:4] == b"%PDF"


def test_build_pdf_contains_contract_type():
    result = build_pdf([{"section": "S1", "summary": "Text"}], "NDA / Confidentiality")
    assert b"NDA" in result
