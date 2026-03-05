from io import BytesIO
from datetime import date

import openpyxl
from openpyxl.styles import Font
from fpdf import FPDF


def _sanitize(text: str) -> str:
    return text.encode("latin-1", errors="replace").decode("latin-1")


def build_excel(results: list[dict]) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contract Review"
    ws.append(["Section", "Summary"])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    for row in results:
        ws.append([row["section"], row["summary"]])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_pdf(results: list[dict], contract_type: str) -> bytes:
    pdf = FPDF()
    pdf.compress = False
    pdf.set_margins(left=20, top=20, right=20)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, _sanitize("Scrutiny"), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", size=9)
    pdf.cell(0, 6, _sanitize(f"Generated: {date.today().isoformat()}"), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "I", 11)
    pdf.cell(0, 8, _sanitize(contract_type), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    for i, row in enumerate(results, 1):
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 7, _sanitize(f"{i}. {row['section']}"), new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", size=10)
        pdf.multi_cell(0, 6, _sanitize(row["summary"]))
        pdf.ln(3)

    return bytes(pdf.output())
